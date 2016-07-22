# -*- coding: utf-8 -*-

from netacad import app
import pymysql
import csv
import sys
import re
from functools import wraps
from flask import request,current_app
from flask.ext.jsonpify import jsonify
import json
from openpyxl import load_workbook
from tools import *

@app.route("/.api/netacad/rank_institution",  methods=['GET'])
@support_jsonp('get_location_rank_institution')
def table_rank_institutions():
    struct=[]

    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT id_institution, nb_student_lvl_1_ccna_discovery + nb_student_lvl_2_ccna_discovery + nb_student_lvl_3_ccna_discovery + nb_student_lvl_4_ccna_discovery + nb_student_lvl_1_ccna_exploration + nb_student_lvl_2_ccna_exploration + nb_student_lvl_3_ccna_exploration + nb_student_lvl_4_ccna_exploration + nb_student_introduction_networks_ccna_routing_switching + nb_student_network_basic_ccna_routing_switching + nb_student_routing_protocols_ccna_routing_switching + nb_student_switching_essentials_ccna_routing_switching + nb_student_scaling_networks_ccna_routing_switching + nb_student_switched_networks_ccna_routing_switching + nb_student_connecting_networks_ccna_routing_switching + nb_student_ccna_security + nb_student_ccnp + nb_student_it_essential AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) ORDER BY total DESC limit 20')
        data = cur.fetchall()

        for row in data :
            cur.execute('SELECT num_id_netacad, name_institution, city_institution  FROM institution WHERE id_institution='+str(row[0])+'')
            info_institution = cur.fetchone()
            struct.append([str(info_institution[0]), str(info_institution[1]), info_institution[2].decode("latin-1"), str(row[1])])

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e
        
        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e


    return 0

@app.route("/.api/netacad/ccna_routing_switching",  methods=['GET'])
@support_jsonp('get_location_exploration')
def ccna_routing_switching():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT SUM(`nb_student_introduction_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_introduction_networks_ccna_routing_switching= cur.fetchone()
        nb_student_introduction_networks_ccna_routing_switching=str(nb_student_introduction_networks_ccna_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_network_basic_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_network_basic_ccna_routing_switching= cur.fetchone()
        nb_student_network_basic_ccna_routing_switching=str(nb_student_network_basic_ccna_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_routing_protocols_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_routing_protocols_ccna_routing_switching= cur.fetchone()
        nb_student_routing_protocols_ccna_routing_switching=str(nb_student_routing_protocols_ccna_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_switching_essentials_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_switching_essentials_ccna_routing_switching= cur.fetchone()
        nb_student_switching_essentials_ccna_routing_switching=str(nb_student_switching_essentials_ccna_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_scaling_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_scaling_networks_ccna_routing_switching= cur.fetchone()
        nb_student_scaling_networks_ccna_routing_switching=str(nb_student_scaling_networks_ccna_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_switched_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_switched_networks_ccna_routing_switching= cur.fetchone()
        nb_student_switched_networks_ccna_routing_switching=str(nb_student_switched_networks_ccna_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_connecting_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_connecting_networks_ccna_routing_switching= cur.fetchone()
        nb_student_connecting_networks_ccna_routing_switching=str(nb_student_connecting_networks_ccna_routing_switching[0])

        struct = [["introduction_networks", nb_student_introduction_networks_ccna_routing_switching],
                  ["network_basic", nb_student_network_basic_ccna_routing_switching],
                  ["routing_protocols", nb_student_routing_protocols_ccna_routing_switching],
                  ["switching_essentials", nb_student_switching_essentials_ccna_routing_switching],
                  ["scaling_networks", nb_student_scaling_networks_ccna_routing_switching],
                  ["switched_networks", nb_student_switched_networks_ccna_routing_switching],
                  ["connecting_networks", nb_student_connecting_networks_ccna_routing_switching]
        ]

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/ccna_exploration",  methods=['GET'])
@support_jsonp('get_location_exploration')
def ccna_exploration():
    struct=[]

    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT SUM(`nb_student_lvl_1_ccna_exploration`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl1_exploration= cur.fetchone()
        nb_student_lvl1_exploration=str(nb_student_lvl1_exploration[0])

        cur.execute('SELECT SUM(`nb_student_lvl_2_ccna_exploration`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl2_exploration= cur.fetchone()
        nb_student_lvl2_exploration=str(nb_student_lvl2_exploration[0])

        cur.execute('SELECT SUM(`nb_student_lvl_3_ccna_exploration`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl3_exploration= cur.fetchone()
        nb_student_lvl3_exploration=str(nb_student_lvl3_exploration[0])

        cur.execute('SELECT SUM(`nb_student_lvl_4_ccna_exploration`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl4_exploration= cur.fetchone()
        nb_student_lvl4_exploration=str(nb_student_lvl4_exploration[0])

        struct = [["level_1", nb_student_lvl1_exploration],
                  ["level_2", nb_student_lvl2_exploration],
                  ["level_3", nb_student_lvl3_exploration],
                  ["level_4", nb_student_lvl4_exploration]
        ]

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/ccna_discovery",  methods=['GET'])
@support_jsonp('get_location_discovery')
def ccna_discovery():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT SUM(`nb_student_lvl_1_ccna_discovery`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl1_discovery=cur.fetchone()
        nb_student_lvl1_discovery=str(nb_student_lvl1_discovery[0])

        cur.execute('SELECT SUM(`nb_student_lvl_2_ccna_discovery`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl2_discovery=cur.fetchone()
        nb_student_lvl2_discovery=str(nb_student_lvl2_discovery[0])

        cur.execute('SELECT SUM(`nb_student_lvl_3_ccna_discovery`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl3_discovery=cur.fetchone()
        nb_student_lvl3_discovery=str(nb_student_lvl3_discovery[0])

        cur.execute('SELECT SUM(`nb_student_lvl_4_ccna_discovery`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_lvl4_discovery=cur.fetchone()
        nb_student_lvl4_discovery=str(nb_student_lvl4_discovery[0])

        struct = [["level_1", nb_student_lvl1_discovery],
                  ["level_2", nb_student_lvl2_discovery],
                  ["level_3", nb_student_lvl3_discovery],
                  ["level_4", nb_student_lvl4_discovery]
        ]

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/academy_students",  methods=['GET'])
@support_jsonp('get_location_student')
def academy_students():
    struct=[]

    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:
        cur.execute('SELECT (SUM(`nb_student_lvl_1_ccna_discovery`)+SUM(`nb_student_lvl_2_ccna_discovery`)+SUM(`nb_student_lvl_3_ccna_discovery`)+SUM(`nb_student_lvl_4_ccna_discovery`)) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_discovery = cur.fetchone()
        nb_student_discovery=str(nb_student_discovery[0])

        cur.execute('SELECT (SUM(`nb_student_lvl_1_ccna_exploration`)+SUM(`nb_student_lvl_2_ccna_exploration`)+SUM(`nb_student_lvl_3_ccna_exploration`)+SUM(`nb_student_lvl_4_ccna_exploration`)) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_exploration = cur.fetchone()
        nb_student_exploration = str(nb_student_exploration[0])


        cur.execute('SELECT (SUM(`nb_student_introduction_networks_ccna_routing_switching`)+SUM(`nb_student_network_basic_ccna_routing_switching`)+SUM(`nb_student_routing_protocols_ccna_routing_switching`)+SUM(`nb_student_switching_essentials_ccna_routing_switching`)+SUM(`nb_student_scaling_networks_ccna_routing_switching`)+SUM(`nb_student_switched_networks_ccna_routing_switching`)+SUM(`nb_student_connecting_networks_ccna_routing_switching`)) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_routing_switching = cur.fetchone()
        nb_student_routing_switching = str(nb_student_routing_switching[0])

        cur.execute('SELECT SUM(`nb_student_ccna_security`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_security= cur.fetchone()
        nb_student_security= str(nb_student_security[0])

        cur.execute('SELECT SUM(`nb_student_ccnp`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_ccnp=cur.fetchone()
        nb_student_ccnp=str(nb_student_ccnp[0])


        cur.execute('SELECT SUM(`nb_student_it_essential`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_student_essential=cur.fetchone()
        nb_student_essential=str(nb_student_essential[0])

        cur.execute('SELECT SUM(nb_student_lvl_1_ccna_discovery + nb_student_lvl_2_ccna_discovery + nb_student_lvl_3_ccna_discovery + nb_student_lvl_4_ccna_discovery + nb_student_lvl_1_ccna_exploration + nb_student_lvl_2_ccna_exploration + nb_student_lvl_3_ccna_exploration + nb_student_lvl_4_ccna_exploration + nb_student_introduction_networks_ccna_routing_switching + nb_student_network_basic_ccna_routing_switching + nb_student_routing_protocols_ccna_routing_switching + nb_student_switching_essentials_ccna_routing_switching + nb_student_scaling_networks_ccna_routing_switching + nb_student_switched_networks_ccna_routing_switching + nb_student_connecting_networks_ccna_routing_switching + nb_student_ccna_security + nb_student_ccnp + nb_student_it_essential) AS total FROM institution, metrics WHERE state_institution != "" AND institution.id_institution=metrics.id_institution AND date_resume IN (SELECT max(date_resume) FROM metrics)')
        nb_students = cur.fetchone()
        nb_students = str(nb_students[0])

        struct = [["Discovery", nb_student_discovery],
                  ["Exploration", nb_student_exploration],
                  ["Routing_Switching", nb_student_routing_switching],
                  ["Security", nb_student_security],
                  ["CCNP", nb_student_ccnp],
                  ["IT Essential", nb_student_essential],
                  ["Students", nb_students]
        ]


        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/academy_ccna",  methods=['GET'])
@support_jsonp('get_location_curriculum_trending')
def academy_ccna():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT COUNT(name_institution) FROM institution')
        nb_institution = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_lvl_1_ccna_discovery != 0 OR nb_student_lvl_2_ccna_discovery != 0 OR nb_student_lvl_3_ccna_discovery != 0 OR nb_student_lvl_4_ccna_discovery != 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_discovery = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_lvl_1_ccna_exploration != 0 OR nb_student_lvl_2_ccna_exploration != 0 OR nb_student_lvl_3_ccna_exploration != 0 OR nb_student_lvl_4_ccna_exploration != 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_exploration = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_introduction_networks_ccna_routing_switching != 0 OR nb_student_network_basic_ccna_routing_switching != 0 OR nb_student_routing_protocols_ccna_routing_switching != 0 OR nb_student_switching_essentials_ccna_routing_switching != 0 OR nb_student_scaling_networks_ccna_routing_switching != 0 OR nb_student_switched_networks_ccna_routing_switching != 0 OR nb_student_connecting_networks_ccna_routing_switching != 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_routing_switching = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_ccna_security != 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_security = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_ccnp != 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_ccnp = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_it_essential != 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_essential = cur.fetchone()

        cur.execute('SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_introduction_networks_ccna_routing_switching = 0 AND nb_student_network_basic_ccna_routing_switching = 0 AND nb_student_routing_protocols_ccna_routing_switching = 0 AND nb_student_switching_essentials_ccna_routing_switching = 0 AND nb_student_scaling_networks_ccna_routing_switching = 0 AND nb_student_switched_networks_ccna_routing_switching = 0 AND nb_student_connecting_networks_ccna_routing_switching = 0 AND nb_student_it_essential = 0 AND nb_student_ccnp = 0 AND nb_student_ccna_security = 0 AND nb_student_lvl_1_ccna_exploration = 0 AND nb_student_lvl_2_ccna_exploration = 0 AND nb_student_lvl_3_ccna_exploration = 0 AND nb_student_lvl_4_ccna_exploration = 0 AND nb_student_lvl_1_ccna_discovery = 0 AND nb_student_lvl_2_ccna_discovery = 0 AND nb_student_lvl_3_ccna_discovery = 0 AND nb_student_lvl_4_ccna_discovery = 0) AND (date_resume IN (SELECT max(date_resume) FROM metrics))')
        nb_no_active = cur.fetchone()


        struct = [["Academies", nb_institution],
                  ["Discovery", nb_discovery],
                  ["Exploration", nb_exploration],
                  ["Routing_Switching", nb_routing_switching],
                  ["Security", nb_security],
                  ["CCNP", nb_ccnp],
                  ["IT Essential", nb_essential],
                  ["No Active", nb_no_active]
        ]


        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/label_location",  methods=['GET'])
@support_jsonp('get_location')
def geolocation():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT name_institution, addr1_institution, city_institution, zip_institution, latitude, longitude FROM institution')
        row = cur.fetchall()
        for line in row:
            #if city exist
            if line[4] != 0 and line[5] != 0:
				 struct.append(['<div class="info_content">'+'<h3>'+line[0].decode('latin-1')+'</h3>'+'<p><b>Address : </b>'+line[1].decode('latin-1')+'</p>'+'<p><b>City : </b>'+line[2].decode('latin-1')+'</p>'+'<p><b>Zip : </b>'+line[3]+' </p>'+'</div>', line[4], line[5]])
			
            """
			if line[3] != "":
                if line[3] == "Paris":
                    cur.execute('SELECT ville_longitude_deg, ville_latitude_deg FROM villes_france WHERE ville_code_postal ="'+line[4]+'" ')
                    search_by_zip = cur.fetchone()

                    if search_by_zip is not None:
                        struct.append(['<div class="info_content">'+'<h3>'+line[0]+'</h3>'+'<p><b>First address : </b>'+line[1]+'</p>'+'<p><b>Second address : </b></p>'+'<p><b>Zip : </b>'+line[4]+' </p>'+'</div>',search_by_zip[1], search_by_zip[0]])

                else:
                    cur.execute('SELECT ville_longitude_deg, ville_latitude_deg FROM villes_france WHERE ville_nom ="'+line[3]+'" ')
                    search_by_city = cur.fetchone()


                 # if result match
                if search_by_city is not None:
                    if line[3] == "Paris":
                        struct.append(['<div class="info_content">'+'<h3>'+line[0]+'</h3>'+'<p><b>First address : </b>'+line[1]+'</p>'+'<p><b>Second address : </b>'+line[2]+'</p>'+'<p><b>City : </b>'+line[3]+' </p>'+'<p><b>Zip : </b>'+line[4]+' </p>'+'</div>',search_by_city[1], search_by_city[0]])

            else:
                #if city doesn't exist but zip exist
                if line[4] != "":
                    cur.execute('SELECT ville_longitude_deg, ville_latitude_deg FROM villes_france WHERE ville_code_postal ="'+line[4]+'" ')
                    search_by_zip = cur.fetchone()

                    if search_by_zip is not None:
                        struct.append(['<div class="info_content">'+'<h3>'+line[0]+'</h3>'+'<p><b>First address : </b>'+line[1]+'</p>'+'<p><b>Second address : </b></p>'+'<p><b>Zip : </b>'+line[4]+' </p>'+'</div>',search_by_zip[1], search_by_zip[0]])

                #if city and zip don't exist
                else:
                    pass
            """

        try :
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e
		
        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e


    return 0

@app.route("/.api/netacad/populate_database_v2")
def populate_database_v2():
    try:

        with open('/var/www/netacad/csv/Academy_Briefing_Book_Details_May_FY15_v1.csv', 'rU') as csvfile :

            try:

                reader = csv.reader(csvfile,  dialect=csv.excel_tab, delimiter=';')

                timeline=1

                for row in reader:

                    if timeline == 2:
                        date_document=row[0]
                        date_document=date_document.replace(',', '')
                        date_document=date_document.split(' ')
                        month_document=date_document[1]
                        year_document=date_document[3]
                        day_document=date_document[2]

                        if months[month_document] != None :
                            date_resume=year_document+"-"+months[month_document]+"-"+day_document
                    #print row[9]
                    #if row[10]=="France":
                    #    print row[10],",",row[0],",",row[1]

                    if timeline >= 7 and row[10] == "France":

                        cnx, cur = ConnectToSqlServer()

                        cur.execute('SELECT num_id_netacad FROM institution WHERE num_id_netacad="'+row[0]+'"')
                        num_netacad = cur.fetchone()

                        if num_netacad is None:
                            cur.execute('INSERT INTO institution (num_id_netacad, name_institution, addr1_institution, addr2_institution, city_institution, state_institution, zip_institution, phone_institution) VALUES ('+row[0]+', "'+row[1]+'", "'+row[13]+'", "'+row[14]+'", "'+row[15]+'", "'+row[16]+'", "'+row[17]+'", "'+row[18]+'")')

                            cur.execute('SELECT id_institution FROM institution WHERE num_id_netacad="'+row[0]+'"')

                            id_institu = cur.fetchone()

                        else:
                            cur.execute('SELECT id_institution FROM institution WHERE num_id_netacad="'+str(num_netacad[0])+'"')
                            id_institu = cur.fetchone()


                        cur.execute('INSERT INTO metrics (bool_cisco_academy, bool_asc_cisco_academy, bool_itc_cisco_academy, account_status_institution_role_attributes, date_current_status_institution_role_attributes, institution_created_date_institution_role_attributes, bool_active_academy, passed_class_rule_active_academy, pass_ramp_up_rule_active_academy, ccna_fqtd, ccnas_fqtd, ccnp_fqtd, ite_fqtd, ccna_fytd, ccnas_fytd, ccnp_fytd, ite_fytd, nb_student_lvl_1_ccna_discovery, nb_student_lvl_2_ccna_discovery, nb_student_lvl_3_ccna_discovery, nb_student_lvl_4_ccna_discovery, nb_student_lvl_1_ccna_exploration, nb_student_lvl_2_ccna_exploration, nb_student_lvl_3_ccna_exploration, nb_student_lvl_4_ccna_exploration, nb_student_introduction_networks_ccna_routing_switching, nb_student_network_basic_ccna_routing_switching, nb_student_routing_protocols_ccna_routing_switching, nb_student_switching_essentials_ccna_routing_switching, nb_student_scaling_networks_ccna_routing_switching, nb_student_switched_networks_ccna_routing_switching, nb_student_connecting_networks_ccna_routing_switching, nb_student_ccna_security, nb_student_ccnp, nb_student_it_essential, year_time, month_time, day_time, date_resume, id_institution, total_students, total_students_female, total_active_instructor, total_active_instructor_female, total_cumul_students_inception, total_cumul_students_inception_female, total_any_graduate_inception, total_any_graduate_inception_female, total_certification_ready_inception, total_certification_ready_inception_female) VALUES ("'+row[2]+'", "'+row[3]+'", "'+row[4]+'", "'+row[25]+'", "'+row[26]+'", "'+row[27]+'", "'+row[28]+'", "'+row[29]+'", "'+row[30]+'", "'+row[40]+'", "'+row[41]+'", "'+row[42]+'", "'+row[43]+'", "'+row[46]+'", "'+row[47]+'", "'+row[48]+'", "'+row[49]+'", "'+row[63]+'", "'+row[64]+'", "'+row[65]+'", "'+row[66]+'", "'+row[68]+'", "'+row[69]+'", "'+row[70]+'", "'+row[71]+'", "'+row[73]+'", "'+row[74]+'", "'+row[76]+'", "'+row[75]+'", "'+row[77]+'", "'+row[78]+'", "'+row[79]+'", "'+row[80]+'", "'+row[81]+'", "'+row[83]+'", "'+str(year_document)+'", "'+str(month_document)+'", "'+str(day_document)+'","'+date_resume+'", "'+str(id_institu[0])+'", "'+row[55]+'", "'+row[56]+'", "'+row[128]+'", "'+row[129]+'", "'+row[130]+'", "'+row[131]+'", "'+row[132]+'", "'+row[133]+'", "'+row[134]+'", "'+row[135]+'")')

                    timeline=timeline+1

            except Exception as e :
                print e

            finally :
                csvfile.close()

    except Exception as e:
        print e

    return 0

@app.route("/.api/netacad/populate_database")
def populate_database():
    try:

        with open('/var/www/netacad/csv/Academy_Briefing_Book_Details_March_FY14_v1.csv', 'rU') as csvfile :

            try:

                reader = csv.reader(csvfile,  dialect=csv.excel_tab, delimiter=';')

                timeline=1

                for row in reader:

                    if timeline == 2:
                        date_document=row[0]
                        date_document=date_document.replace(',', '')
                        date_document=date_document.split(' ')
                        month_document=date_document[1]
                        year_document=date_document[3]
                        day_document=date_document[2]

                        if months[month_document] != None :
                            date_resume=year_document+"-"+months[month_document]+"-"+day_document
                    print row[9]
                    if row[9]=="France":
                        print row[9],",",row[0],",",row[1]

                    if timeline >= 7 and row[9] == "France":

                        cnx, cur = ConnectToSqlServer()

                        cur.execute('SELECT num_id_netacad FROM institution WHERE num_id_netacad="'+row[0]+'"')
                        num_netacad = cur.fetchone()

                        if num_netacad is None:
                            cur.execute('INSERT INTO institution (num_id_netacad, name_institution, addr1_institution, addr2_institution, city_institution, state_institution, zip_institution, phone_institution) VALUES ('+row[0]+', "'+row[1]+'", "'+row[13]+'", "'+row[14]+'", "'+row[15]+'", "'+row[16]+'", "'+row[17]+'", "'+row[18]+'")')

                            cur.execute('SELECT id_institution FROM institution WHERE num_id_netacad="'+row[0]+'"')

                            id_institu = cur.fetchone()

                        else:
                            cur.execute('SELECT id_institution FROM institution WHERE num_id_netacad="'+str(num_netacad[0])+'"')
                            id_institu = cur.fetchone()


                        cur.execute('INSERT INTO metrics (bool_cisco_academy, bool_asc_cisco_academy, bool_itc_cisco_academy, account_status_institution_role_attributes, date_current_status_institution_role_attributes, institution_created_date_institution_role_attributes, bool_active_academy, passed_class_rule_active_academy, pass_ramp_up_rule_active_academy, ccna_fqtd, ccnas_fqtd, ccnp_fqtd, ite_fqtd, ccna_fytd, ccnas_fytd, ccnp_fytd, ite_fytd, nb_student_lvl_1_ccna_discovery, nb_student_lvl_2_ccna_discovery, nb_student_lvl_3_ccna_discovery, nb_student_lvl_4_ccna_discovery, nb_student_lvl_1_ccna_exploration, nb_student_lvl_2_ccna_exploration, nb_student_lvl_3_ccna_exploration, nb_student_lvl_4_ccna_exploration, nb_student_introduction_networks_ccna_routing_switching, nb_student_network_basic_ccna_routing_switching, nb_student_routing_protocols_ccna_routing_switching, nb_student_switching_essentials_ccna_routing_switching, nb_student_scaling_networks_ccna_routing_switching, nb_student_switched_networks_ccna_routing_switching, nb_student_connecting_networks_ccna_routing_switching, nb_student_ccna_security, nb_student_ccnp, nb_student_it_essential, year_time, month_time, day_time, date_resume, id_institution, total_students, total_students_female, total_active_instructor, total_active_instructor_female, total_cumul_students_inception, total_cumul_students_inception_female, total_any_graduate_inception, total_any_graduate_inception_female, total_certification_ready_inception, total_certification_ready_inception_female) VALUES ("'+row[2]+'", "'+row[3]+'", "'+row[4]+'", "'+row[25]+'", "'+row[26]+'", "'+row[27]+'", "'+row[28]+'", "'+row[29]+'", "'+row[30]+'", "'+row[40]+'", "'+row[41]+'", "'+row[42]+'", "'+row[43]+'", "'+row[46]+'", "'+row[47]+'", "'+row[48]+'", "'+row[49]+'", "'+row[63]+'", "'+row[64]+'", "'+row[65]+'", "'+row[66]+'", "'+row[68]+'", "'+row[69]+'", "'+row[70]+'", "'+row[71]+'", "'+row[73]+'", "'+row[74]+'", "'+row[76]+'", "'+row[75]+'", "'+row[77]+'", "'+row[78]+'", "'+row[79]+'", "'+row[80]+'", "'+row[81]+'", "'+row[83]+'", "'+str(year_document)+'", "'+str(month_document)+'", "'+str(day_document)+'","'+date_resume+'", "'+str(id_institu[0])+'", "'+row[55]+'", "'+row[56]+'", "'+row[128]+'", "'+row[129]+'", "'+row[130]+'", "'+row[131]+'", "'+row[132]+'", "'+row[133]+'", "'+row[134]+'", "'+row[135]+'")')

                    timeline=timeline+1

            except Exception as e :
                print e

            finally :
                csvfile.close()

    except Exception as e:
        print e

    return 0

@app.route("/.api/netacad/populate_database_xls")
def populate_database_xls():

    try:
        wb = load_workbook(filename="/var/www/netacad/csv/Academy_Briefing_Book_Details_September_FY15_v1.xlsx")
        ws = wb.active
        print(wb.get_sheet_names())
        timeline=0
        for row in ws.iter_rows():
            for cell in row:

                if timeline == 2 :
                    date_document= str(ws.cell(row = timeline, column = 1).value)
                    date_document=date_document.replace(',', '')
                    date_document=date_document.split(' ')
                    month_document=date_document[1]
                    year_document=date_document[3]
                    day_document=date_document[2]

                    #print "month_document :"+month_document
                    #print "year : "+year_document
                    #print "day : "+day_document

                    if months[month_document] != None :
                        date_resume=year_document+"-"+months[month_document]+"-"+day_document


                if timeline >= 7 and ws.cell(row = timeline, column = 10).value == 'France':

                    #print "num_id_netcad : "+str(ws.cell(row = timeline, column = 1).value)
                    #print "name_institution : "+str(ws.cell(row = timeline, column = 2).value)
                    #print "addr1 : "+str(ws.cell(row = timeline, column = 14).value)
                    #print "city : "+str(ws.cell(row = timeline, column = 16).value)
                    print "level type : "+str(ws.cell(row = timeline, column = 37).value)

                    #cnx, cur = ConnectToSqlServer()

                    #cur.execute('SELECT num_id_netacad FROM institution WHERE num_id_netacad="'+row[0]+'"')
                    #num_netacad = cur.fetchone()

                    #if num_netacad is None:
                        #cur.execute('INSERT INTO institution (num_id_netacad, name_institution, addr1_institution, addr2_institution, city_institution, state_institution, zip_institution, phone_institution) VALUES ('+row[0]+', "'+row[1]+'", "'+row[13]+'", "'+row[14]+'", "'+row[15]+'", "'+row[16]+'", "'+row[17]+'", "'+row[18]+'")')

                        #cur.execute('SELECT id_institution FROM institution WHERE num_id_netacad="'+row[0]+'"')

                    #    id_institu = cur.fetchone()

                    #else:
                        #print ""
                        #cur.execute('SELECT id_institution FROM institution WHERE num_id_netacad="'+str(num_netacad[0])+'"')
                    #    id_institu = cur.fetchone()


                    #cur.execute('INSERT INTO metrics (bool_cisco_academy, bool_asc_cisco_academy, bool_itc_cisco_academy, account_status_institution_role_attributes, date_current_status_institution_role_attributes, institution_created_date_institution_role_attributes, bool_active_academy, passed_class_rule_active_academy, pass_ramp_up_rule_active_academy, ccna_fqtd, ccnas_fqtd, ccnp_fqtd, ite_fqtd, ccna_fytd, ccnas_fytd, ccnp_fytd, ite_fytd, nb_student_lvl_1_ccna_discovery, nb_student_lvl_2_ccna_discovery, nb_student_lvl_3_ccna_discovery, nb_student_lvl_4_ccna_discovery, nb_student_lvl_1_ccna_exploration, nb_student_lvl_2_ccna_exploration, nb_student_lvl_3_ccna_exploration, nb_student_lvl_4_ccna_exploration, nb_student_introduction_networks_ccna_routing_switching, nb_student_network_basic_ccna_routing_switching, nb_student_routing_protocols_ccna_routing_switching, nb_student_switching_essentials_ccna_routing_switching, nb_student_scaling_networks_ccna_routing_switching, nb_student_switched_networks_ccna_routing_switching, nb_student_connecting_networks_ccna_routing_switching, nb_student_ccna_security, nb_student_ccnp, nb_student_it_essential, year_time, month_time, day_time, date_resume, id_institution) VALUES ("'+row[2]+'", "'+row[3]+'", "'+row[4]+'", "'+row[25]+'", "'+row[26]+'", "'+row[27]+'", "'+row[28]+'", "'+row[29]+'", "'+row[30]+'", "'+row[40]+'", "'+row[41]+'", "'+row[42]+'", "'+row[43]+'", "'+row[46]+'", "'+row[47]+'", "'+row[48]+'", "'+row[49]+'", "'+row[63]+'", "'+row[64]+'", "'+row[65]+'", "'+row[66]+'", "'+row[68]+'", "'+row[69]+'", "'+row[70]+'", "'+row[71]+'", "'+row[73]+'", "'+row[74]+'", "'+row[76]+'", "'+row[75]+'", "'+row[77]+'", "'+row[78]+'", "'+row[79]+'", "'+row[80]+'", "'+row[81]+'", "'+row[83]+'", "'+str(year_document)+'", "'+str(month_document)+'", "'+str(day_document)+'","'+date_resume+'", "'+str(id_institu[0])+'")')

                timeline=timeline+1



                #if ws.cell(row = i, column = 10).value == 'France' :


                   # print "plop0"
                    #print str(ws.cell(row = i, column = 17).value)


    except Exception as e:
        print e

    return 0

@app.route("/.api/netacad/geo_loc_institutions",  methods=['GET'])
@support_jsonp('get_location_geo_loc_institutions')
def geo_loc_insitutions():
    struct=[]

    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT state_institution, COUNT(*) AS total FROM institution GROUP BY state_institution ORDER BY total DESC')
        data = cur.fetchall()

        for row in data :
            if row[0] != "" :
                if row[0] == "Rhone-Alpes" :
                     struct.append(["Rhône-Alpes", str(row[1])])

                else :
                    struct.append([row[0].decode("latin-1"), str(row[1])])

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/geo_loc_students",  methods=['GET'])
@support_jsonp('get_location_geo_loc_students')
def geo_loc_students():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:

        cur.execute('SELECT state_institution, SUM(nb_student_lvl_1_ccna_discovery + nb_student_lvl_2_ccna_discovery + nb_student_lvl_3_ccna_discovery + nb_student_lvl_4_ccna_discovery + nb_student_lvl_1_ccna_exploration + nb_student_lvl_2_ccna_exploration + nb_student_lvl_3_ccna_exploration + nb_student_lvl_4_ccna_exploration + nb_student_introduction_networks_ccna_routing_switching + nb_student_network_basic_ccna_routing_switching + nb_student_routing_protocols_ccna_routing_switching + nb_student_switching_essentials_ccna_routing_switching + nb_student_scaling_networks_ccna_routing_switching + nb_student_switched_networks_ccna_routing_switching + nb_student_connecting_networks_ccna_routing_switching + nb_student_ccna_security + nb_student_ccnp + nb_student_it_essential) AS total FROM institution, metrics WHERE state_institution != "" AND institution.id_institution=metrics.id_institution AND date_resume IN (SELECT max(date_resume) FROM metrics) GROUP BY state_institution ORDER BY total DESC')
        data = cur.fetchall()

        for row in data :
            if row[0] != "" :
                if row[0] == "Rhone-Alpes" :
                    struct.append(["Rhône-Alpes", str(row[1])])
                else :
                    struct.append([row[0].decode("latin-1"), str(row[1])])

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/geo_click_region",  methods=['GET'])
@support_jsonp('get_location_geo_click_region')
def geo_click_region():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:
        message=request.args.get('message').encode('UTF-8')
        print "#################### test 1"
        print type(message)
        print "message_click : "+message
        print "#################### test 2"
        cur.execute('SELECT id_institution FROM institution WHERE state_institution like "'+message+'"')
        data = cur.fetchall()
        print "#################### test 3"
        #((points,),)=cur.fetchall()
        #print "data : "+points
        current_student_discovery=0
        current_student_essential=0
        current_student_security=0
        current_student_exploration=0
        current_student_routing_switching=0
        current_student_ccnp=0
        #print cur.fetchall()

        for row in data:

            cmd="SELECT (SUM(`nb_student_lvl_1_ccna_discovery`)+SUM(`nb_student_lvl_2_ccna_discovery`)+SUM(`nb_student_lvl_3_ccna_discovery`)+SUM(`nb_student_lvl_4_ccna_discovery`)) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) AND id_institution="+str(row[0])
            cur.execute(cmd)
            nb_student_discovery = cur.fetchone()
            nb_student_discovery=int(nb_student_discovery[0])

            cmd='SELECT (SUM(`nb_student_lvl_1_ccna_exploration`)+SUM(`nb_student_lvl_2_ccna_exploration`)+SUM(`nb_student_lvl_3_ccna_exploration`)+SUM(`nb_student_lvl_4_ccna_exploration`)) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) AND id_institution='+str(row[0])
            cur.execute(cmd)
            nb_student_exploration = cur.fetchone()
            nb_student_exploration = int(nb_student_exploration[0])

            cmd='SELECT (SUM(`nb_student_introduction_networks_ccna_routing_switching`)+SUM(`nb_student_network_basic_ccna_routing_switching`)+SUM(`nb_student_routing_protocols_ccna_routing_switching`)+SUM(`nb_student_switching_essentials_ccna_routing_switching`)+SUM(`nb_student_scaling_networks_ccna_routing_switching`)+SUM(`nb_student_switched_networks_ccna_routing_switching`)+SUM(`nb_student_connecting_networks_ccna_routing_switching`)) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) AND id_institution='+str(row[0])
            cur.execute(cmd)
            nb_student_routing_switching = cur.fetchone()
            nb_student_routing_switching = int(nb_student_routing_switching[0])

            cmd='SELECT SUM(`nb_student_ccna_security`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) AND id_institution='+str(row[0])
            cur.execute(cmd)
            nb_student_security= cur.fetchone()
            nb_student_security= int(nb_student_security[0])

            cmd='SELECT SUM(`nb_student_ccnp`) AS total FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) AND id_institution='+str(row[0])
            cur.execute(cmd)
            nb_student_ccnp=cur.fetchone()
            nb_student_ccnp=int(nb_student_ccnp[0])

            cmd='SELECT SUM(`nb_student_it_essential`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics) AND id_institution='+str(row[0])
            cur.execute(cmd)
            nb_student_essential=cur.fetchone()
            nb_student_essential=int(nb_student_essential[0])

            current_student_discovery=current_student_discovery+nb_student_discovery
            current_student_exploration=current_student_exploration+nb_student_exploration
            current_student_routing_switching=current_student_routing_switching+nb_student_routing_switching
            current_student_security=current_student_security+nb_student_security
            current_student_ccnp=current_student_ccnp+nb_student_ccnp
            current_student_essential=current_student_essential+nb_student_essential


        struct =[["Discovery", current_student_discovery],
                    ["Exploration", current_student_exploration],
                    ["Routing", current_student_routing_switching],
                    ["Security", current_student_security],
                    ["CCNP", current_student_ccnp],
                    ["IT Essential", current_student_essential]
        ]


        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e



@app.route("/.api/netacad/gender",  methods=['GET'])
@support_jsonp('get_location_gender')
def academy_gender():
    struct=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:
        cmd="SELECT SUM(`total_students`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_students = cur.fetchone()
        total_students = str(total_students[0])

        cmd="SELECT SUM(`total_students_female`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_students_female = cur.fetchone()
        total_students_female=int(total_students_female[0])

        cmd="SELECT SUM(`total_active_instructor`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_active_instructor = cur.fetchone()
        total_active_instructor = str(total_active_instructor[0])

        cmd="SELECT SUM(`total_active_instructor_female`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_active_instructor_female = cur.fetchone()
        total_active_instructor_female = str(total_active_instructor_female[0])

        cmd="SELECT SUM(`total_cumul_students_inception`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_cumul_students_inception = cur.fetchone()
        total_cumul_students_inception = str(total_cumul_students_inception[0])

        cmd="SELECT SUM(`total_cumul_students_inception_female`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_cumul_students_inception_female = cur.fetchone()
        total_cumul_students_inception_female = str(total_cumul_students_inception_female[0])

        cmd="SELECT SUM(`total_any_graduate_inception`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_any_graduate_inception = cur.fetchone()
        total_any_graduate_inception = str(total_any_graduate_inception[0])

        cmd="SELECT SUM(`total_any_graduate_inception_female`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_any_graduate_inception_female = cur.fetchone()
        total_any_graduate_inception_female = str(total_any_graduate_inception_female[0])

        cmd="SELECT SUM(`total_certification_ready_inception`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_certification_ready_inception = cur.fetchone()
        total_certification_ready_inception = str(total_certification_ready_inception[0])

        cmd="SELECT SUM(`total_certification_ready_inception_female`) FROM metrics WHERE date_resume IN (SELECT max(date_resume) FROM metrics)"
        cur.execute(cmd)
        total_certification_ready_inception_female = cur.fetchone()
        total_certification_ready_inception_female = str(total_certification_ready_inception_female[0])


        struct = [["Total", total_students],
                  ["Female", total_students_female],
                  ["Total", total_active_instructor],
                  ["Female", total_active_instructor_female],
                  ["Total", total_cumul_students_inception],
                  ["Female", total_cumul_students_inception_female],
                  ["Total", total_any_graduate_inception],
                  ["Female", total_any_graduate_inception_female],
                  ["Total", total_certification_ready_inception],
                  ["Female", total_certification_ready_inception_female]
        ]
        #print struct

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct))

    except Exception as e:
        print e

@app.route("/.api/netacad/academy_ccna_trending",  methods=['GET'])
@support_jsonp('get_location_curriculum_trending')
def academy_ccna_trending():
    struct_gobal=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:
        i=5
        cur.execute('select date_resume from metrics GROUP BY date_resume')
        date = cur.fetchall()

        for row in date :


            # Academies metrics
            cmd="SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_lvl_1_ccna_discovery != 0 OR nb_student_lvl_2_ccna_discovery != 0 OR nb_student_lvl_3_ccna_discovery != 0 OR nb_student_lvl_4_ccna_discovery != 0) AND date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_discovery = cur.fetchone()
            nb_discovery = int(nb_discovery[0])

            cmd="SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_lvl_1_ccna_exploration != 0 OR nb_student_lvl_2_ccna_exploration != 0 OR nb_student_lvl_3_ccna_exploration != 0 OR nb_student_lvl_4_ccna_exploration != 0) AND date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_exploration = cur.fetchone()
            nb_exploration = int(nb_exploration[0])

            cmd="SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_introduction_networks_ccna_routing_switching != 0 OR nb_student_network_basic_ccna_routing_switching != 0 OR nb_student_routing_protocols_ccna_routing_switching != 0 OR nb_student_switching_essentials_ccna_routing_switching != 0 OR nb_student_scaling_networks_ccna_routing_switching != 0 OR nb_student_switched_networks_ccna_routing_switching != 0 OR nb_student_connecting_networks_ccna_routing_switching != 0) AND date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_routing_switching = cur.fetchone()
            nb_routing_switching = int(nb_routing_switching[0])

            cmd="SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_ccna_security != 0) AND date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_security = cur.fetchone()
            nb_security = int(nb_security[0])

            cmd="SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_ccnp != 0) AND date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_ccnp = cur.fetchone()
            nb_ccnp = int(nb_ccnp[0])

            cmd="SELECT COUNT(DISTINCT(id_institution)) FROM metrics WHERE (nb_student_it_essential != 0) AND date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_essential = cur.fetchone()
            nb_essential = int(nb_essential[0])

            # Students metrics
            cmd="SELECT (SUM(`nb_student_lvl_1_ccna_discovery`)+SUM(`nb_student_lvl_2_ccna_discovery`)+SUM(`nb_student_lvl_3_ccna_discovery`)+SUM(`nb_student_lvl_4_ccna_discovery`)) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_discovery = cur.fetchone()
            nb_student_discovery=str(nb_student_discovery[0])

            cmd="SELECT (SUM(`nb_student_lvl_1_ccna_exploration`)+SUM(`nb_student_lvl_2_ccna_exploration`)+SUM(`nb_student_lvl_3_ccna_exploration`)+SUM(`nb_student_lvl_4_ccna_exploration`)) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_exploration = cur.fetchone()
            nb_student_exploration = str(nb_student_exploration[0])

            cmd="SELECT (SUM(`nb_student_introduction_networks_ccna_routing_switching`)+SUM(`nb_student_network_basic_ccna_routing_switching`)+SUM(`nb_student_routing_protocols_ccna_routing_switching`)+SUM(`nb_student_switching_essentials_ccna_routing_switching`)+SUM(`nb_student_scaling_networks_ccna_routing_switching`)+SUM(`nb_student_switched_networks_ccna_routing_switching`)+SUM(`nb_student_connecting_networks_ccna_routing_switching`)) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_routing_switching = cur.fetchone()
            nb_student_routing_switching = str(nb_student_routing_switching[0])

            cmd="SELECT SUM(`nb_student_ccna_security`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_security= cur.fetchone()
            nb_student_security= str(nb_student_security[0])

            cmd="SELECT SUM(`nb_student_ccnp`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_ccnp=cur.fetchone()
            nb_student_ccnp=str(nb_student_ccnp[0])

            cmd="SELECT SUM(`nb_student_it_essential`) FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_essential=cur.fetchone()
            nb_student_essential=str(nb_student_essential[0])


            struct = [
                    ["date", str(row[0])],
                  ["Discovery", nb_discovery],
                  ["Exploration", nb_exploration],
                  ["Routing_Switching", nb_routing_switching],
                  ["Security", nb_security],
                  ["CCNP", nb_ccnp],
                  ["IT Essential", nb_essential],
                  ["Discovery", nb_student_discovery],
                  ["Exploration", nb_student_exploration],
                  ["Routing_Switching", nb_student_routing_switching],
                  ["Security", nb_student_security],
                  ["CCNP", nb_student_ccnp],
                  ["IT Essential", nb_student_essential]
            ]
            struct_gobal.append(struct)

            i=i-1
            if i==0 :
                break

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct_gobal))

    except Exception as e:
        print e


@app.route("/.api/netacad/split_curriculum_trending",  methods=['GET'])
@support_jsonp('get_location_split_curriculum_trending')
def split_curriculum_trending():
    struct_gobal=[]


    try:
        cnx, cur = ConnectToSqlServer()
    except Exception as e:
        print e

    try:
        i=5
        cur.execute('select date_resume from metrics GROUP BY date_resume')
        date = cur.fetchall()

        for row in date :

            cmd="SELECT SUM(`nb_student_lvl_1_ccna_discovery`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl1_discovery=cur.fetchone()
            nb_student_lvl1_discovery=str(nb_student_lvl1_discovery[0])

            cmd="SELECT SUM(`nb_student_lvl_1_ccna_discovery`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl2_discovery=cur.fetchone()
            nb_student_lvl2_discovery=str(nb_student_lvl2_discovery[0])

            cmd="SELECT SUM(`nb_student_lvl_3_ccna_discovery`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl3_discovery=cur.fetchone()
            nb_student_lvl3_discovery=str(nb_student_lvl3_discovery[0])

            cmd="SELECT SUM(`nb_student_lvl_4_ccna_discovery`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl4_discovery=cur.fetchone()
            nb_student_lvl4_discovery=str(nb_student_lvl4_discovery[0])

            cmd="SELECT SUM(`nb_student_lvl_1_ccna_exploration`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl1_exploration=cur.fetchone()
            nb_student_lvl1_exploration=str(nb_student_lvl1_exploration[0])

            cmd="SELECT SUM(`nb_student_lvl_2_ccna_exploration`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl2_exploration=cur.fetchone()
            nb_student_lvl2_exploration=str(nb_student_lvl2_exploration[0])

            cmd="SELECT SUM(`nb_student_lvl_3_ccna_exploration`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl3_exploration=cur.fetchone()
            nb_student_lvl3_exploration=str(nb_student_lvl3_exploration[0])

            cmd="SELECT SUM(`nb_student_lvl_4_ccna_exploration`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_lvl4_exploration=cur.fetchone()
            nb_student_lvl4_exploration=str(nb_student_lvl4_exploration[0])

            cmd="SELECT SUM(`nb_student_introduction_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_introduction_networks_ccna_routing_switching= cur.fetchone()
            nb_student_introduction_networks_ccna_routing_switching=str(nb_student_introduction_networks_ccna_routing_switching[0])

            cmd="SELECT SUM(`nb_student_network_basic_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_network_basic_ccna_routing_switching= cur.fetchone()
            nb_student_network_basic_ccna_routing_switching=str(nb_student_network_basic_ccna_routing_switching[0])

            cmd="SELECT SUM(`nb_student_routing_protocols_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_routing_protocols_ccna_routing_switching= cur.fetchone()
            nb_student_routing_protocols_ccna_routing_switching=str(nb_student_routing_protocols_ccna_routing_switching[0])

            cmd="SELECT SUM(`nb_student_switching_essentials_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_switching_essentials_ccna_routing_switching= cur.fetchone()
            nb_student_switching_essentials_ccna_routing_switching=str(nb_student_switching_essentials_ccna_routing_switching[0])

            cmd="SELECT SUM(`nb_student_scaling_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_scaling_networks_ccna_routing_switching= cur.fetchone()
            nb_student_scaling_networks_ccna_routing_switching=str(nb_student_scaling_networks_ccna_routing_switching[0])

            cmd="SELECT SUM(`nb_student_switched_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_switched_networks_ccna_routing_switching= cur.fetchone()
            nb_student_switched_networks_ccna_routing_switching=str(nb_student_switched_networks_ccna_routing_switching[0])

            cmd="SELECT SUM(`nb_student_connecting_networks_ccna_routing_switching`) AS total FROM metrics WHERE date_resume='"+str(row[0])+"'"
            cur.execute(cmd)
            nb_student_connecting_networks_ccna_routing_switching= cur.fetchone()
            nb_student_connecting_networks_ccna_routing_switching=str(nb_student_connecting_networks_ccna_routing_switching[0])

            struct = [["date", str(row[0])],
                ["level_1", nb_student_lvl1_discovery],
                ["level_2", nb_student_lvl2_discovery],
                ["level_3", nb_student_lvl3_discovery],
                ["level_4", nb_student_lvl4_discovery],
                ["level_1", nb_student_lvl1_exploration],
                ["level_2", nb_student_lvl2_exploration],
                ["level_3", nb_student_lvl3_exploration],
                ["level_4", nb_student_lvl4_exploration],
                ["introduction_networks", nb_student_introduction_networks_ccna_routing_switching],
                ["network_basic", nb_student_network_basic_ccna_routing_switching],
                ["routing_protocols", nb_student_routing_protocols_ccna_routing_switching],
                ["switching_essentials", nb_student_switching_essentials_ccna_routing_switching],
                ["scaling_networks", nb_student_scaling_networks_ccna_routing_switching],
                ["switched_networks", nb_student_switched_networks_ccna_routing_switching],
                ["connecting_networks", nb_student_connecting_networks_ccna_routing_switching]
            ]


            struct_gobal.append(struct)

            i=i-1
            if i==0 :
                break

        try:
            CloseToSqlServer(cnx, cur)
        except Exception as e :
            print e

        return jsonify(summary=json.dumps(struct_gobal))

    except Exception as e:
        print e
